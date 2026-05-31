"""
v4.20–v4.25 스캔 검출 종목 정량적 근거 스냅샷 → Excel (타임라인 고정·GUI 무관).

산출 경로: outputs/evidences/Scan_Evidence_{code}_{anchor}.xlsx
"""
from __future__ import annotations

import math
import os
import re
from dataclasses import dataclass
from typing import Any, Mapping, Sequence

import numpy as np
import pandas as pd

from src.filters import (
    PULLBACK_DISPARITY20_LOCK_PCT,
    PULLBACK_DISPARITY5_LOCK_PCT,
    pass_disparity_lock,
)

EVIDENCE_OUTPUT_DIR = os.path.join("outputs", "evidences")
OHLC_LAYER = "당일 가격 (OHLC) ★"
META_LAYER = "종목 메타 정보"

EvidenceRow = tuple[str, str, str, str, str, str]


@dataclass(frozen=True)
class ScanEvidenceSnapshot:
    """스캔 시점 고정 문자열 스냅샷 — 이후 GUI·시장 변경과 무관."""

    code: str
    name: str
    anchor_date: str
    listing_market: str
    rows: tuple[EvidenceRow, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "name": self.name,
            "anchor_date": self.anchor_date,
            "listing_market": self.listing_market,
            "rows": [list(r) for r in self.rows],
        }

    @classmethod
    def from_dict(cls, data: Mapping[str, Any]) -> "ScanEvidenceSnapshot":
        raw_rows = data.get("rows") or []
        rows: list[EvidenceRow] = []
        for el in raw_rows:
            if isinstance(el, (list, tuple)) and len(el) >= 6:
                rows.append(
                    (
                        str(el[0]),
                        str(el[1]),
                        str(el[2]),
                        str(el[3]),
                        str(el[4]),
                        str(el[5]),
                    )
                )
        return cls(
            code=str(data.get("code", "")).zfill(6),
            name=str(data.get("name", "")),
            anchor_date=str(data.get("anchor_date", ""))[:10],
            listing_market=str(data.get("listing_market", "KOSPI")),
            rows=tuple(rows),
        )

    def with_display_name(self, name: str) -> "ScanEvidenceSnapshot":
        """종목명 행만 갱신 — OHLC·수치 스냅샷은 불변."""
        nm = str(name or "").strip() or self.code
        patched: list[EvidenceRow] = []
        label = f"{nm} ({self.code})"
        for layer, metric, val, thr, verdict, guide in self.rows:
            if metric == "대상 종목명 / 코드":
                patched.append((layer, metric, label, thr, verdict, guide))
            else:
                patched.append((layer, metric, val, thr, verdict, guide))
        return ScanEvidenceSnapshot(
            code=self.code,
            name=nm,
            anchor_date=self.anchor_date,
            listing_market=self.listing_market,
            rows=tuple(patched),
        )


def _market_guide(listing_market: str) -> str:
    m = str(listing_market or "").strip().upper()
    if m == "KOSDAQ":
        return "코스닥 상장 유니버스"
    if m == "KOSPI":
        return "코스피 상장 유니버스"
    return "코스피·코스닥 상장 유니버스"


def _finite(x: object) -> float | None:
    try:
        v = float(x)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
    return v if math.isfinite(v) else None


def _fmt_krw_eok_label(value_krw: float | None) -> str:
    v = _finite(value_krw)
    if v is None or v <= 0:
        return "-"
    eok = int(round(v / 1e8))
    return f"{eok:,d} 억 원"


def _fmt_price_krw(value: float | None) -> str:
    """원본 OHLC — 반올림 정수 원 문자열 (조작·재계산 없음)."""
    v = _finite(value)
    if v is None or v <= 0:
        return "-"
    return f"{int(round(v)):,d} 원"


def _verdict_pass_fail(ok: bool) -> str:
    return "PASS" if ok else "FAIL 🚨"


def _verdict_info() -> str:
    return "INFO"


def _disparity_pct(close: float | None, ma: float | None) -> float | None:
    c, m = _finite(close), _finite(ma)
    if c is None or m is None or m <= 0:
        return None
    return c / m * 100.0


def _safe_sheet_title(name: str, code: str) -> str:
    base = re.sub(r'[\\/*?:\[\]]', "_", str(name or code).strip())[:20]
    return f"{base}_근거"[:31] or str(code).zfill(6)


def build_scan_evidence_from_metrics(
    *,
    code: str,
    name: str,
    anchor_date: str,
    listing_market: str,
    open_t0: float | None,
    high_t0: float | None,
    low_t0: float | None,
    close_t0: float | None,
    market_cap_krw: float | None,
    trade_amount_krw: float | None,
    min_liquidity_market_cap_krw: float,
    min_liquidity_trade_amount_krw: float,
    volume_burst_multiple: float,
    vol_shrink_limit: float,
    prev_vol: float | None,
    vol_ma20_prior: float | None,
    prev_open: float | None,
    prev_close: float | None,
    ma5: float | None,
    ma20: float | None,
    ma60: float | None,
    ma120: float | None,
    today_vol: float | None,
) -> ScanEvidenceSnapshot:
    """벌크 merged row·단일 OHLCV 공통 — t0 OHLC 원본 + 수치 문자열 고정."""
    code6 = str(code).strip().zfill(6)
    nm = str(name or "").strip() or code6
    anchor = str(anchor_date).strip()[:10]
    lm = str(listing_market or "KOSPI").strip().upper()

    ot, ht, lt, ct = (
        _finite(open_t0),
        _finite(high_t0),
        _finite(low_t0),
        _finite(close_t0),
    )

    min_cap = float(min_liquidity_market_cap_krw)
    min_trd = float(min_liquidity_trade_amount_krw)
    burst_req = float(volume_burst_multiple)
    shrink_req = float(vol_shrink_limit)

    cap_ok = _finite(market_cap_krw) is not None and float(market_cap_krw) >= min_cap  # type: ignore[arg-type]
    trd_ok = _finite(trade_amount_krw) is not None and float(trade_amount_krw) >= min_trd  # type: ignore[arg-type]

    pv, vma = _finite(prev_vol), _finite(vol_ma20_prior)
    vol_mult = (pv / vma) if pv and vma and vma > 0 else None
    vol_mult_ok = vol_mult is not None and vol_mult >= burst_req

    po, pc = _finite(prev_open), _finite(prev_close)
    prev_yang = pc is not None and po is not None and pc > po

    m5v, m20v = _finite(ma5), _finite(ma20)
    disp5 = _disparity_pct(ct, m5v)
    disp20 = _disparity_pct(ct, m20v)
    disp_ok = (
        ct is not None
        and m5v is not None
        and m20v is not None
        and pass_disparity_lock(ct, m5v, m20v)
    )
    disp5_ok = disp5 is not None and disp5 <= PULLBACK_DISPARITY5_LOCK_PCT
    disp20_ok = disp20 is not None and disp20 <= PULLBACK_DISPARITY20_LOCK_PCT

    tv, pvv = _finite(today_vol), _finite(prev_vol)
    shrink_ratio = (tv / pvv) if tv is not None and pvv and pvv > 0 else None
    shrink_ok = shrink_ratio is not None and shrink_ratio <= shrink_req

    m60, m120 = _finite(ma60), _finite(ma120)
    trend_ok = (
        ct is not None
        and m60 is not None
        and m120 is not None
        and ct > m60
        and ct > m120
        and m60 > m120
    )

    cap_thr = f"{int(round(min_cap / 1e8)):,d}억 원 이상"
    trd_thr = f"{int(round(min_trd / 1e8)):,d}억 원 이상"

    rows: list[EvidenceRow] = [
        (
            META_LAYER,
            "조회 기준일 (t0)",
            anchor,
            "사용자 선택 앵커일",
            _verdict_info(),
            "스캔이 가동된 타겟 거래일 세션",
        ),
        (
            META_LAYER,
            "대상 종목명 / 코드",
            f"{nm} ({code6})",
            "-",
            _verdict_info(),
            _market_guide(lm),
        ),
        (
            OHLC_LAYER,
            "당일 시가 (Open)",
            _fmt_price_krw(ot),
            "-",
            _verdict_info(),
            "교차 검증용 실제 시가",
        ),
        (
            OHLC_LAYER,
            "당일 고가 (High)",
            _fmt_price_krw(ht),
            "-",
            _verdict_info(),
            "교차 검증용 실제 고가",
        ),
        (
            OHLC_LAYER,
            "당일 저가 (Low)",
            _fmt_price_krw(lt),
            "-",
            _verdict_info(),
            "교차 검증용 실제 저가 (5분봉·루머 진위 확인용)",
        ),
        (
            OHLC_LAYER,
            "당일 종가 (Close)",
            _fmt_price_krw(ct),
            "-",
            _verdict_info(),
            "교차 검증용 실제 종가",
        ),
        (
            "Pass 0 : 유동성 필터",
            "당일 시가총액",
            _fmt_krw_eok_label(market_cap_krw),
            cap_thr,
            _verdict_pass_fail(cap_ok),
            "소형 품절주 및 잡주 1차 차단",
        ),
        (
            "Pass 0 : 유동성 필터",
            "당일 거래대금",
            _fmt_krw_eok_label(trade_amount_krw),
            trd_thr,
            _verdict_pass_fail(trd_ok),
            "최소한의 장중 유동성 마지노선",
        ),
        (
            "Pass 1 : 세력 개입 수급",
            "t-1 거래량 스파이크 배수",
            f"{vol_mult:.2f} 배" if vol_mult is not None else "-",
            f"{burst_req:g} 배 이상",
            _verdict_pass_fail(vol_mult_ok),
            "평균 거래량 대비 세력 유입 증명",
        ),
        (
            "Pass 1 : 세력 개입 수급",
            "t-1 캔들 양봉 여부",
            "양봉 마감 (종가 > 시가)" if prev_yang else "음봉 또는 데이터 없음",
            "무조건 양봉 필수",
            _verdict_pass_fail(prev_yang),
            "음봉 설거지형 노이즈 차단",
        ),
        (
            "Pass 2 : 눌림목 & 이격도",
            "5일 이격도 (Disparity 5)",
            f"{disp5:.2f} %" if disp5 is not None else "-",
            f"{PULLBACK_DISPARITY5_LOCK_PCT:.1f}% 이하 (락)",
            _verdict_pass_fail(disp5_ok),
            (
                "5일선 한참 위 과열 상태 (진입 금지)"
                if not disp5_ok
                else "5일 이격도 락 통과"
            ),
        ),
        (
            "Pass 2 : 눌림목 & 이격도",
            "20일 이격도 (Disparity 20)",
            f"{disp20:.2f} %" if disp20 is not None else "-",
            f"{PULLBACK_DISPARITY20_LOCK_PCT:.1f}% 이하 (락)",
            _verdict_pass_fail(disp20_ok),
            (
                "20일선 상단 붕 뜬 자리 추격매수 차단"
                if not disp20_ok
                else "20일 이격도 락 통과"
            ),
        ),
        (
            "Pass 2 : 눌림목 & 이격도",
            "20일 이동평균선 (MA20)",
            _fmt_price_krw(m20v),
            "-",
            _verdict_info(),
            "저가·이격도 교차 검증용 MA20 스냅샷",
        ),
        (
            "Pass 3 : 거래량 동결",
            "당일 거래량 감소 비율",
            (
                f"{shrink_ratio:.2f} ({int(round(shrink_ratio * 100))}%)"
                if shrink_ratio is not None
                else "-"
            ),
            f"{shrink_req:g} 이하 (감소)",
            _verdict_pass_fail(shrink_ok),
            "세력 이탈 없는 눌림 거래량 수렴",
        ),
        (
            "Pass 4 : Perfect Trend",
            "60일선 / 120일선 위치",
            "MA60 > MA120 (정배열)" if trend_ok else "역배열 또는 미충족",
            "배열성 강제 락",
            _verdict_pass_fail(trend_ok),
            "장기 우상향 추세 담보 구조",
        ),
    ]
    _ = disp_ok  # 스캔 게이트와 동기 — 리포트는 항목별 판정으로 표기
    return ScanEvidenceSnapshot(
        code=code6,
        name=nm,
        anchor_date=anchor,
        listing_market=lm,
        rows=tuple(rows),
    )


def build_scan_evidence_from_ohlcv(
    df: pd.DataFrame,
    *,
    code: str,
    name: str,
    anchor_date: str,
    listing_market: str,
    market_cap_krw: float | None,
    trade_amount_krw: float | None,
    min_liquidity_market_cap_krw: float,
    min_liquidity_trade_amount_krw: float,
    volume_burst_multiple: float,
    vol_shrink_limit: float,
) -> ScanEvidenceSnapshot | None:
    """폴백 단일 종목 OHLCV — t0 봉 Open/High/Low/Close 원본 그대로."""
    from src.data_loader import ensure_datetime_index
    from src.filters import PULLBACK_MIN_OHLCV_BARS

    if df is None or df.empty:
        return None
    work = ensure_datetime_index(df.copy()).sort_index()
    if len(work) < PULLBACK_MIN_OHLCV_BARS:
        return None

    vol = pd.to_numeric(work["Volume"], errors="coerce")
    close = pd.to_numeric(work["Close"], errors="coerce")
    opn = pd.to_numeric(work["Open"], errors="coerce")
    high = pd.to_numeric(work["High"], errors="coerce")
    low = pd.to_numeric(work["Low"], errors="coerce")

    return build_scan_evidence_from_metrics(
        code=code,
        name=name,
        anchor_date=anchor_date,
        listing_market=listing_market,
        open_t0=float(opn.iloc[-1]),
        high_t0=float(high.iloc[-1]),
        low_t0=float(low.iloc[-1]),
        close_t0=float(close.iloc[-1]),
        market_cap_krw=market_cap_krw,
        trade_amount_krw=trade_amount_krw,
        min_liquidity_market_cap_krw=min_liquidity_market_cap_krw,
        min_liquidity_trade_amount_krw=min_liquidity_trade_amount_krw,
        volume_burst_multiple=volume_burst_multiple,
        vol_shrink_limit=vol_shrink_limit,
        prev_vol=float(vol.iloc[-2]),
        vol_ma20_prior=float(vol.iloc[-22:-2].mean()),
        prev_open=float(opn.iloc[-2]),
        prev_close=float(close.iloc[-2]),
        ma5=float(close.iloc[-5:].mean()),
        ma20=float(close.iloc[-20:].mean()),
        ma60=float(close.iloc[-60:].mean()),
        ma120=float(close.iloc[-120:].mean()),
        today_vol=float(vol.iloc[-1]),
    )


def _apply_evidence_workbook_style(ws) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    navy = PatternFill("solid", fgColor="1E3A5F")
    ohlc_fill = PatternFill("solid", fgColor="E8EEF5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1E3A5F")
    subtitle_font = Font(size=10, color="555555")
    zebra_a = PatternFill("solid", fgColor="F7F9FC")
    zebra_b = PatternFill("solid", fgColor="FFFFFF")
    pass_fill = PatternFill("solid", fgColor="DFF5E4")
    fail_fill = PatternFill("solid", fgColor="FDE8E8")
    info_fill = PatternFill("solid", fgColor="EEF2F7")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 38

    ws["B2"] = "💡 주도주 눌림목 스캐너 검출 근거 리포트 (Snapshot v4.25)"
    ws["B2"].font = title_font
    ws["B3"] = (
        "조회 시점 t0 OHLC 원본·수치를 고정(Snapshot) 기록 — "
        "GUI·시장 변경과 무관한 무결성 데이터."
    )
    ws["B3"].font = subtitle_font

    hdr_row = 5
    headers = (
        "파이프라인 레이어",
        "세부 검증 지표",
        "조회 스냅샷 수치",
        "시스템 통과 기준",
        "판정 결과",
        "트레이더 분석 가이드",
    )
    for col, label in enumerate(headers, start=2):
        cell = ws.cell(row=hdr_row, column=col, value=label)
        cell.fill = navy
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(hdr_row + 1, ws.max_row + 1):
        layer = str(ws.cell(row=r, column=2).value or "").strip()
        verdict = str(ws.cell(row=r, column=6).value or "").strip()
        stripe = zebra_a if (r - hdr_row) % 2 == 1 else zebra_b
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if OHLC_LAYER in layer:
                cell.fill = ohlc_fill
            elif c == 6:
                if verdict.startswith("FAIL"):
                    cell.fill = fail_fill
                elif verdict == "PASS":
                    cell.fill = pass_fill
                elif verdict == "INFO":
                    cell.fill = info_fill
                else:
                    cell.fill = stripe
            else:
                cell.fill = stripe

    ws.freeze_panes = "B6"


def _write_evidence_sheet(ws, snap: ScanEvidenceSnapshot) -> None:
    ws.title = _safe_sheet_title(snap.name, snap.code)
    for i, row in enumerate(snap.rows, start=6):
        for j, val in enumerate(row, start=2):
            ws.cell(row=i, column=j, value=val)
    _apply_evidence_workbook_style(ws)


def generate_evidence_snapshot(
    snap: ScanEvidenceSnapshot,
    *,
    output_dir: str = EVIDENCE_OUTPUT_DIR,
) -> str:
    """단일 종목 evidence xlsx — 스냅샷 문자열만 기록."""
    from openpyxl import Workbook

    os.makedirs(output_dir, exist_ok=True)
    save_path = os.path.join(
        output_dir,
        f"Scan_Evidence_{snap.code}_{snap.anchor_date}.xlsx",
    )
    wb = Workbook()
    ws = wb.active
    _write_evidence_sheet(ws, snap)
    wb.save(save_path)
    return save_path


def export_scan_evidence_snapshots(
    snapshots: Sequence[ScanEvidenceSnapshot],
    *,
    output_dir: str = EVIDENCE_OUTPUT_DIR,
) -> list[str]:
    """검출 종목별 독립 xlsx (GUI 리프레시·시장 변경과 무관)."""
    paths: list[str] = []
    for snap in snapshots:
        paths.append(generate_evidence_snapshot(snap, output_dir=output_dir))
    return paths


def build_scan_evidence_from_bulk_row(
    row: pd.Series,
    *,
    code: str,
    name: str,
    anchor_date: str,
    listing_market: str,
    market_cap_krw: float | None,
    trade_amount_krw: float | None,
    min_liquidity_market_cap_krw: float,
    min_liquidity_trade_amount_krw: float,
    volume_burst_multiple: float,
    vol_shrink_limit: float,
) -> ScanEvidenceSnapshot:
    """벌크 scan merged row — t0 OHLC 컬럼 원본 그대로."""

    def _col(key: str) -> float | None:
        return _finite(row.get(key))

    return build_scan_evidence_from_metrics(
        code=code,
        name=name,
        anchor_date=anchor_date,
        listing_market=listing_market,
        open_t0=_col("Open_t0"),
        high_t0=_col("High_t0"),
        low_t0=_col("Low_t0"),
        close_t0=_col("Close_t0"),
        market_cap_krw=market_cap_krw,
        trade_amount_krw=trade_amount_krw,
        min_liquidity_market_cap_krw=min_liquidity_market_cap_krw,
        min_liquidity_trade_amount_krw=min_liquidity_trade_amount_krw,
        volume_burst_multiple=volume_burst_multiple,
        vol_shrink_limit=vol_shrink_limit,
        prev_vol=_col("prev_vol"),
        vol_ma20_prior=_col("vol_ma20_strictly_prior"),
        prev_open=_col("Prev_open"),
        prev_close=_col("Prev_close"),
        ma5=_col("MA5"),
        ma20=_col("MA20"),
        ma60=_col("MA60"),
        ma120=_col("MA120"),
        today_vol=_col("today_vol"),
    )
