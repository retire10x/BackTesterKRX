"""
v4.60: 백테스트 총괄 요약 + 상세 매매 로그 → Excel (outputs/backtests/).
"""
from __future__ import annotations

import os
from typing import Any, Mapping, Sequence

BACKTEST_OUTPUT_DIR = os.path.join("outputs", "backtests")


def _display_profit_factor(pf: float) -> float:
    if pf != pf:  # NaN
        return 0.0
    if pf == float("inf"):
        return 999.99
    return float(pf)


def export_backtest_evidence(
    ticker_code: str,
    ticker_name: str,
    start_date: str,
    end_date: str,
    metrics_dict: Mapping[str, Any],
    trade_history: Sequence[Mapping[str, Any]],
) -> str:
    """
    v4.60: 백테스트 총괄 요약 및 상세 매매 로그 무결성 고정 엑셀 생성.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    code6 = str(ticker_code or "").strip().zfill(6)
    end_s = str(end_date or "").strip()[:10]

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest_Evidence"
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="맑은 고딕", size=12, bold=True, color="000000")
    font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="맑은 고딕", size=10, bold=False)
    font_fail = Font(name="맑은 고딕", size=10, bold=True, color="9C0006")
    font_pass = Font(name="맑은 고딕", size=10, bold=True, color="006100")

    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    fill_pass_pastel = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_fail_pastel = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_thin = Border(
        left=Side(style="thin", color="D5DBDB"),
        right=Side(style="thin", color="D5DBDB"),
        top=Side(style="thin", color="D5DBDB"),
        bottom=Side(style="thin", color="D5DBDB"),
    )

    total_trades = int(metrics_dict.get("total_trades", 0) or 0)
    win_rate = float(metrics_dict.get("win_rate", 0.0) or 0.0)
    profit_factor = _display_profit_factor(float(metrics_dict.get("profit_factor", 0.0) or 0.0))

    ws.append([])
    ws.append(["", "💡 백테스트 전략 검증 근거 리포트 (Snapshot v4.60)"])
    ws.cell(row=2, column=2).font = font_title
    ws.append(["", "조회 구간 내 발생한 모든 진입/청산 타점의 1원 단위 원본 데이터입니다."])
    ws.cell(row=3, column=2).font = Font(
        name="맑은 고딕", size=9, italic=True, color="7F8C8D"
    )
    ws.append([])

    summary_headers = [
        "파이프라인 레이어",
        "세부 검증 지표",
        "조회 스냅샷 수치",
        "전략 통과 기준",
        "판정 결과",
        "트레이더 분석 가이드",
    ]
    ws.append([""] + summary_headers)
    for col in range(2, 8):
        cell = ws.cell(row=5, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    metadata = [
        [
            "백테스트 메타",
            "대상 종목명 / 코드",
            f"{ticker_name} ({code6})",
            "-",
            "INFO",
            "타겟 검증 우량주",
        ],
        [
            "백테스트 메타",
            "테스트 구동 기간",
            f"{start_date} ~ {end_date}",
            "최근 6개월 고정",
            "INFO",
            "v4.50 동적 버퍼 엔진 가동 구간",
        ],
        [
            "성과 측정 지표",
            "총 매매 횟수",
            f"{total_trades} 회",
            "-",
            "INFO",
            "샘플 모수 통계치",
        ],
        [
            "성과 측정 지표",
            "최종 승률 (Win Rate)",
            f"{win_rate:.2f} %",
            "60.0% 이상 권장",
            "PASS" if win_rate >= 60 else "FAIL",
            "단기 기술적 반등 확률 우위",
        ],
        [
            "성과 측정 지표",
            "평균 손익비 (PF)",
            f"{profit_factor:.2f}",
            "1.50 이상 필수",
            "PASS" if profit_factor >= 1.5 else "FAIL",
            "장기적 복리 우상향 증명 지표",
        ],
    ]

    for row_data in metadata:
        ws.append([""] + row_data)
        curr_row = ws.max_row
        for col in range(2, 8):
            cell = ws.cell(row=curr_row, column=col)
            cell.font = font_body
            cell.border = border_thin
            if col in (2, 3, 5, 6):
                cell.alignment = align_center
            if col == 4:
                cell.alignment = align_right
            if cell.value in ("PASS", "FAIL"):
                cell.fill = fill_pass_pastel if cell.value == "PASS" else fill_fail_pastel
                cell.font = font_pass if cell.value == "PASS" else font_fail

    ws.append([])
    ws.append([])
    trade_headers = [
        "매매 회차",
        "진입일 (t0)",
        "진입가 (종가 매수)",
        "청산일 (t+1)",
        "청산가 (시가 매도)",
        "개별 수익률",
        "매매 판정",
    ]
    ws.append([""] + trade_headers)
    t_header_row = ws.max_row
    for col in range(2, 9):
        cell = ws.cell(row=t_header_row, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for idx, trade in enumerate(trade_history, 1):
        pnl = float(trade.get("pnl_ratio", 0.0) or 0.0)
        pnl_str = f"{pnl:.2f} %"
        state = "익절 🟢" if pnl > 0 else "손절 🚨"
        entry_px = int(round(float(trade.get("entry_price", 0) or 0)))
        exit_px = int(round(float(trade.get("exit_price", 0) or 0)))
        row_data = [
            f"{idx}회차",
            str(trade.get("entry_date", "")),
            f"{entry_px:,} 원",
            str(trade.get("exit_date", "")),
            f"{exit_px:,} 원",
            pnl_str,
            state,
        ]
        ws.append([""] + row_data)
        curr_row = ws.max_row

        is_zebra = idx % 2 == 0
        for col in range(2, 9):
            cell = ws.cell(row=curr_row, column=col)
            cell.font = font_body
            cell.border = border_thin
            cell.alignment = align_center if col in (2, 3, 5, 7, 8) else align_right
            if is_zebra:
                cell.fill = fill_zebra
            if col == 8:
                cell.fill = fill_pass_pastel if "익절" in state else fill_fail_pastel
                cell.font = font_pass if "익절" in state else font_fail

    for col_idx in range(2, 9):
        max_len = 12
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    os.makedirs(BACKTEST_OUTPUT_DIR, exist_ok=True)
    filename = os.path.join(
        BACKTEST_OUTPUT_DIR, f"Backtest_Evidence_{code6}_{end_s}.xlsx"
    )
    wb.save(filename)
    print(f"[DEBUG] v4.60 백테스트 근거 저장 완료 -> {filename}")
    return filename
